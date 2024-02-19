from openpyxl import Workbook, load_workbook
from multiprocessing import Pool, Process, Manager

Parents = ["NZ05", "YM03"]

chromIDmap = {}
for i in range(7):
    for j in ["A", "B", "D"]:
        chromIDmap["chr" + str(i+1) + j] = len(chromIDmap)

with open("02.data.merged.hmp.txt") as f:
    lines = f.readlines()
data = [
    lines[0].replace("\n", "").split("\t")
]
for line in lines[1:]:
    ele = line.replace("\n", "").split("\t")
    ele[3] = float(ele[3])
    data.append(ele)

wb = Workbook()

# sheet 1
ws = wb.active # 新工作簿默认包含一个sheet。使用wb.active进入活动工作表
ws.title = "GeneralInfo"
tdata = [
    [4, "!Mapping Population Type (see remarks above)"],
    [1, "!Mapping Function (1 for Kosambi; 2 for Haldane; 3 for Morgan)"],
    [1, "!Marker Space Type (1 for intervals; 2 for positions)"],
    [len(data)-1, "!Number of Markers"],
    [len(data[0])-11, "!Size of the mapping population"]
]
for line in tdata:
    ws.append(line)

# sheet2
wb.create_sheet('Parents')
ws = wb['Parents']
for MarkerIdx in range(1, len(data)):
    ws.append([data[MarkerIdx][0], "TT", "GG"])

# sheet3
wb.create_sheet('Genotype')
ws = wb['Genotype']
for MarkerIdx in range(1, len(data)):
    tdata = [data[MarkerIdx][0]]
    for SampleIdx in range(11, len(data[0])):
        if data[MarkerIdx][SampleIdx] == "T":
            tdata.append("TT")
        elif data[MarkerIdx][SampleIdx] == "G":
            tdata.append("GG")
        elif data[MarkerIdx][SampleIdx] == "K":
            tdata.append("GT")
        else:
            tdata.append("NN")
    ws.append(tdata)

# sheet4
wb.create_sheet('Anchor')
ws = wb['Anchor']
for MarkerIdx in range(1, len(data)):
    ws.append([data[MarkerIdx][0], chromIDmap[data[MarkerIdx][2]]])

wb.save("04.hmp2excel.icimapping.xlsx")